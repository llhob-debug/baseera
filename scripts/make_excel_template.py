from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

# ====== المسارات ======
BASE = os.path.dirname(os.path.dirname(__file__))  # frontend/
LOGO_PATH = os.path.join(BASE, "public", "brand", "logo.png")

# عدّل اسم ملف التمبلت إذا كان مختلف عندك:
INPUT_XLSX = os.path.join(BASE, "basira_excel_import_ready.xlsx")
OUTPUT_XLSX = os.path.join(BASE, "basira-template.xlsx")

if not os.path.exists(INPUT_XLSX):
    raise FileNotFoundError(f"لم أجد ملف التمبلت: {INPUT_XLSX}")

if not os.path.exists(LOGO_PATH):
    raise FileNotFoundError(f"لم أجد الشعار: {LOGO_PATH}")

wb = load_workbook(INPUT_XLSX)

# نضمن اسم الورقة data
ws = wb.active
ws.title = "data"

# ====== تجهيز رأس الصفحة ======
ws.insert_rows(1, 4)  # مساحة للشعار والعنوان
ws.merge_cells("A1:E1")
ws["A1"] = "نموذج بصيرة لإدخال البيانات المالية"
ws["A1"].font = Font(size=14, bold=True)
ws["A1"].alignment = Alignment(horizontal="right", vertical="center")

# إدراج الشعار
img = XLImage(LOGO_PATH)
img.width = 240
img.height = 56
ws.add_image(img, "A2")

# تحسين عرض الأعمدة
widths = [18, 14, 14, 22, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# تثبيت الصف العلوي (بعد الهيدر) على حسب مكان العناوين
ws.freeze_panes = "A6"

wb.save(OUTPUT_XLSX)
print("✅ تم إنشاء ملف Excel النهائي:", OUTPUT_XLSX)
